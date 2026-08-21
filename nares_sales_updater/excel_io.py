"""Lettura file Excel (xlsx via openpyxl, xls via xlrd) come lista di dict."""
from __future__ import annotations

import datetime
from pathlib import Path

import xlrd

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


class ExcelReadError(RuntimeError):
    pass


def _normalize(value):
    if isinstance(value, datetime.time):
        return value.isoformat()
    return value


def read_excel(path: Path, sheet: str | None = None) -> list[dict]:
    """Legge un file xlsx/xls e restituisce le righe (dalla 2 in poi) come dict
    con chiavi = intestazioni di riga 1."""
    path = Path(path)
    if not path.exists():
        raise ExcelReadError(f"File non trovato: {path}")
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(path, sheet)
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(path, sheet)
    raise ExcelReadError(f"Estensione non supportata: {path} ({suffix})")


def _read_xlsx(path: Path, sheet: str | None) -> list[dict]:
    if openpyxl is None:  # pragma: no cover
        raise ExcelReadError("openpyxl non installato")
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
    except KeyError as exc:
        raise ExcelReadError(f"Foglio '{sheet}' non trovato in {path.name}; disponibili: {wb.sheetnames}") from exc
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for raw in rows[1:]:
        row = {}
        for i, header in enumerate(headers):
            if header == "":
                continue
            value = raw[i] if i < len(raw) else None
            row[header] = _normalize(value)
        data.append(row)
    return data


def _read_xls(path: Path, sheet: str | None) -> list[dict]:
    wb = xlrd.open_workbook(str(path))
    ws = None
    if sheet:
        for s in wb.sheets():
            if s.name == sheet:
                ws = s
                break
        if ws is None:
            raise ExcelReadError(f"Foglio '{sheet}' non trovato in {path.name}; disponibili: {[s.name for s in wb.sheets()]}")
    else:
        ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return []
    headers = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
    data = []
    for r in range(1, ws.nrows):
        row = {}
        for c, header in enumerate(headers):
            if header == "":
                continue
            row[header] = _normalize(ws.cell_value(r, c))
        data.append(row)
    return data


def read_headers(path: Path, sheet: str | None = None) -> list[str]:
    """Solo intestazioni (riga 1), per la validazione del mapping."""
    suffix = Path(path).suffix.lower()
    if suffix == ".xls":
        wb = xlrd.open_workbook(str(path))
        ws = next((s for s in wb.sheets() if sheet is None or s.name == sheet), None)
        if ws is None:
            raise ExcelReadError(f"Foglio '{sheet}' non trovato")
        return [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
    if openpyxl is None:  # pragma: no cover
        raise ExcelReadError("openpyxl non installato")
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
    except KeyError as exc:
        raise ExcelReadError(f"Foglio '{sheet}' non trovato") from exc
    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    wb.close()
    return headers


def write_cleaned_excel(path: Path, columns: list[str], rows: list[dict]) -> Path:
    """Scrive i dati puliti in un xlsx (per ispezione; il caricamento usa il DB)."""
    if openpyxl is None:  # pragma: no cover
        raise ExcelReadError("openpyxl non installato")
    import datetime as _dt

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cleaned"
    ws.append(list(columns))
    for row in rows:
        values = []
        for col in columns:
            value = row.get(col)
            if isinstance(value, _dt.datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, _dt.date):
                value = value.strftime("%Y-%m-%d")
            values.append(value)
        ws.append(values)
    wb.save(path)
    wb.close()
    return path
